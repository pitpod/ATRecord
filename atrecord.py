# -*- coding: utf-8 -*-
import os
import sys
# import sqlite3
import pandas as pd
import datetime as dt
import openpyxl
import calendar
import chardet
# import locale

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# from PyQt5 import QtCore
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QMessageBox
from PyQt5.QtWidgets import QApplication, QTableWidgetItem
# from PyQt5.QtGui import QIcon, QColor, QPainter

from datetime import datetime
from datetime import date
from datetime import timedelta

from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.pagesizes import A4, portrait
from reportlab.lib.units import mm
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from PyQt5.QtPrintSupport import QPrintDialog, QPrinter

from Ui_atrecord import Ui_MainWindow


class Application(QMainWindow):
    def __init__(self, parent=None):
        super(Application, self).__init__(parent)
        self.excel_date = SerialData()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.pushButton.clicked.connect(self.csv_read)
        self.ui.pushButton_2.clicked.connect(lambda:self.excel_write(self.df))
        self.ui.pushButton_3.clicked.connect(lambda:self.pdf_write(self.df))

        self.ui.action_F.triggered.connect(lambda: self.csv_read())
        self.ui.action_F.triggered.connect(lambda: self.excel_write())
        self.ui.action_P.triggered.connect(lambda: self.pdf_write())
        self.ui.action_B.triggered.connect(lambda: self.version("バージョン：1.0.0"))

    def version(self, ms_text):
        msgBox = QMessageBox()
        msgBox.setText(ms_text)
        msgBox.setIcon(QMessageBox.Icon.Information)
        msgBox.setStandardButtons(QMessageBox.Ok)
        msgBox.exec_()

    def csv_read(self):
        fname = QFileDialog.getOpenFileName(self, 'Open file', '')
        filepath = fname[0]
        if filepath == "":
            return "break"
        # self.ui.listWidget.addItem(filepath)
        with open(filepath, 'rb') as f:
            c = f.read()
            result = chardet.detect(c)
        if result['encoding'] == 'SHIFT_JIS':
            file_encoding = 'CP932'
        else:
            file_encoding = result['encoding']
        self.df = pd.read_csv(filepath, header=0, skiprows=5, skipfooter=3, encoding=file_encoding)
        # self.df_pdf = pd.read_csv(filepath, header=None, skiprows=5, encoding=file_encoding)
        self.ui.listWidget.addItem("\t".join(self.df.columns.values))

        for index, data in self.df.iterrows():
            list1 = data.to_list()
            map1 = map(str, list1)
            list2 = list(map1)
            self.ui.listWidget.addItem("\t".join(list2))

    def excel_write(self, datadf):
        """EXCELファイルを作成

        Args:
            datadf (データフレーム): 勤怠データ
        """
        last = len(datadf) - 1
        start_date = datadf.iloc[0,0]
        start_date_array = start_date.split("-")
        start_year = start_date_array[0]
        start_month = start_date_array[1]
        start_month_last = calendar.monthrange(int(start_year), int(start_month))[1]

        end_date = datadf.iloc[last,0]
        end_date_array = end_date.split("-")
        end_month = end_date_array[1]
        at_record_range = start_month_last - 20 + 20
        self.excel_serial = self.excel_date.excel_serial(f"{start_year}/{start_month}/21")

        excel_pass = os.path.dirname(__file__)
        self.wb = openpyxl.load_workbook(f"{excel_pass}/TimeSheet.xlsx")
        data_range = f"タイムシート期間:{start_month}月21日~{end_month}月20日"
        ws = self.wb.worksheets[0]
        ws["A1"] = data_range
        data_row = 0
        for i in range(at_record_range):
            date_col = self.excel_serial + i
            date_row = i + 9
            date_key = str(self.excel_date.excel_date(date_col).date())
            d_t = pd.to_datetime(date_key)
            df_array = self.df.query('日付 == @date_key')
            if self.df.query('日付 == @date_key').empty == False:
                start_time = df_array.at[df_array.index[0],"出勤"]
                end_time = df_array.at[df_array.index[0],"退勤"]
                rest_time = df_array.at[df_array.index[0],"休憩"]
                remarks = df_array.at[df_array.index[0],"備考"]
                if remarks == "NaN":
                    remarks = ""
                ws.cell(row=date_row, column=2, value=start_time)
                ws.cell(row=date_row, column=3, value=end_time)
                ws.cell(row=date_row, column=4, value=rest_time)
                ws.cell(row=date_row, column=6, value=remarks)
                data_row += 1

            ws.cell(row=date_row, column=1, value=date_col)
        year_month = d_t.strftime('%Y%m')
        save_fname = QFileDialog.getSaveFileName(self, 'Save File', f'{year_month}.xlsx')
        save_filepass = save_fname[0]
        if save_filepass == "":
            return "break"
        self.wb.save(save_filepass)
        self.wb.close()

        msgBox = QMessageBox()
        msgBox.setText("書き出しました")
        msgBox.setIcon(QMessageBox.Icon.Information)
        msgBox.setStandardButtons(QMessageBox.Ok)
        msgBox.exec_()

    def pdf_write(self, datadf):

        last = len(datadf) - 1
        start_date = datadf.iloc[0,0]
        start_date_array = start_date.split("-")
        start_year = start_date_array[0]
        start_month = start_date_array[1]
        start_month_last = calendar.monthrange(int(start_year), int(start_month))[1]

        end_date = datadf.iloc[last,0]
        end_date_array = end_date.split("-")
        end_month = end_date_array[1]
        at_record_range = start_month_last - 20 + 20
        self.dt = date(int(start_year), int(start_month), 21)
        date_list = []
        columns1 = ["日付", "出勤", "退勤", "休憩", "就業時間", "備考"]
        total_wt = 0

        for i in range(at_record_range):
            date_key = f'{self.dt}'
            d_t = pd.to_datetime(date_key)
            week_day = self.get_day_of_week_jp(d_t)
            # date_time = d_t.strftime('%#d日')
            date_time = d_t.strftime('%d日')
            df_array = datadf.query('日付 == @date_key')
            if datadf.query('日付 == @date_key').empty == False:
                s_t = pd.to_datetime(df_array.at[df_array.index[0],"出勤"])
                # start_time = s_t.strftime('%#H:%M')
                start_time = s_t.strftime('%H:%M')
                e_t = pd.to_datetime(df_array.at[df_array.index[0],"退勤"])
                # end_time = e_t.strftime('%#H:%M')
                end_time = e_t.strftime('%H:%M')
                rest_time = df_array.at[df_array.index[0],"休憩"]
                rt = timedelta(minutes=int(rest_time))
                wt = pd.to_datetime(df_array.at[df_array.index[0],'退勤']) - pd.to_datetime(df_array.at[df_array.index[0],'出勤'])
                wt = wt.seconds - rt.seconds
                total_wt = total_wt + wt
                work_time = self.get_h_m_s(wt)
                remarks = df_array.at[df_array.index[0],"備考"]
                if str(remarks) == "nan":
                    remarks = ""
                list_row = [f'{date_time}{week_day}', f'{start_time}', f'{end_time}', f'{rest_time}', work_time , remarks]
            else:
                list_row = [f'{date_time}{week_day}', "", "", "", "", ""]

            date_list.append(list_row)
            self.dt = self.dt + timedelta(days=1)
        
        year_month = d_t.strftime('%Y%m')
        datadf_full = pd.DataFrame(data=date_list, columns=columns1)

        datadf_full['承認'] = ""
        datadf_pdf = datadf_full.T.reset_index().T.values.tolist()

        # 縦型A4のCanvasを準備
        save_pdfname = QFileDialog.getSaveFileName(self, 'Save File', f'{year_month}.pdf')
        save_pdfpass = save_pdfname[0]
        if save_pdfpass == "":
            return "break"
        cv = canvas.Canvas(save_pdfpass, pagesize=portrait(A4))
        #cv.setLineWidth(5)
        #cv.setDash([5, 5, 5])
        # フォント登録
        pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
        table = Table(datadf_pdf, colWidths=(25*mm, 20*mm, 20*mm, 20*mm, 20*mm, 70*mm, 20*mm), rowHeights=7*mm)
        table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'HeiseiKakuGo-W5', 12), # フォント
            ('BOX', (0, 0), (-1, -1), 1, colors.black),       # 罫線
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            ('LINEBEFORE', (1, 0), (-1, 0), 0.5, colors.black),
            ('INNERGRID', (0, 1), (-1, -1), 0.5 , colors.black) ,
            ('ALIGN', (0,0), (-1, 0), 'CENTER')
        ]))
        origin = 297 - 50 - 7*(at_record_range + 1)
        table.wrapOn(cv, 10*mm, origin*mm) # table位置
        table.drawOn(cv, 10*mm, origin*mm)

        table2_data = (['就業日数 :', f'{len(datadf)}'],['交通費 :', ''])
        table2 = Table(table2_data, colWidths=(25*mm, 40*mm), rowHeights=7*mm)
        table2.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'HeiseiKakuGo-W5', 12), # フォント
            ('BOX', (0, 0), (-1, -1), 1, colors.black),       # 罫線
            ("LINEABOVE", (0, 1), (-1, 1), 1, colors.black),
            ('ALIGN', (0,0), (-1, -1), 'LEFT')
        ]))

        table2.wrapOn(cv, 140*mm, 250*mm) # table位置
        table2.drawOn(cv, 140*mm, 250*mm)

        # 線の太さ
        cv.setLineWidth(1)
        # 線を描画(始点x、始点y、終点x、終点y)
        # フォントサイズ定義
        font_size = 12
        cv.setFont('HeiseiKakuGo-W5', font_size)

        cv.line(65*mm, 286*mm, 145*mm, 286*mm)
        sheet_title = f"タイムシート期間:{start_month}月21日~{end_month}月20日"
        cv.drawString(65*mm, (286 + 1)*mm, sheet_title)

        line_start = 278
        line_height = 7
        all_time = '就業時間(全時間):'
        line_title = ['氏名：松本幸治',
                      '就業先会社名：所属会社名：株式会社ながさきUUカンパニー',
                      '就業先担当者：',
                      f'{all_time}{self.get_h_m_s(total_wt)}']
        for i in range(4):
            cv.line(10*mm, line_start*mm, 130*mm, line_start*mm)
            cv.drawString(10*mm, (line_start + 1)*mm, f'{line_title[i]}')
            line_start = line_start - line_height

        # 保存
        cv.showPage()
        cv.save()

        msgBox = QMessageBox()
        msgBox.setText("書き出しました")
        msgBox.setIcon(QMessageBox.Icon.Information)
        msgBox.setStandardButtons(QMessageBox.Ok)
        msgBox.exec_()

    def get_h_m_s(self, tds):
        """時:分形式を取得

        Args:
            tds (int): 秒数

        Returns:
            (str) : '時:分'
        """
        m, s = divmod(tds, 60)
        h, m = divmod(m, 60)
        # return h, m, s
        h = str(h)
        m = str(f'{m:02}')
        hm = f'{h}:{m}'
        return hm

    def get_day_of_week_jp(self, dt):
        """ 曜日の日本語表示
        Args:
            dt (timestamp): デートタイム
        """
        w_list = ['(月)', '(火)', '(水)', '(木)', '(金)', '(土)', '(日)']
        return(w_list[dt.weekday()])

    def print_dialog(self, printer_obj):
        dialog = QPrintDialog(printer_obj, None)
        dialog.setWindowTitle("Print Document")
        return dialog

class SerialData():
    """ EXCELの日付とシリアル値の変換および
        ウィークナンバーの日本語への変換
    """
    def excel_date(self, date1):
        """ 日付をEXCELのシリアル値に変換
        Args:
            date1 (int): 日数
        """
        print(type(date1))
        temp = datetime(1899, 12, 30)  # Note, not 31st Dec but 30th!
        return(temp + timedelta(days=date1))

    def excel_serial(self, date2):
        """日付テキストを受けてエクセルのシリアル値を返す
        Args:
            date2 (TEXT): YYYY/MM/DD 形式の日付
        """
        date2_sep = date2.split('/')
        day_count = datetime(int(date2_sep[0]), int(date2_sep[1]), int(date2_sep[2]))
        temp = datetime(1899, 12, 30)  # Note, not 31st Dec but 30th!
        return((day_count - temp).days)

    def week_day(self, wd):
        """ウイークナンバーを日本語に変換
        Args:
            wd (int): ウィークナンバー

        Returns:
            string: 曜日の日本語名
        """
        wn_wd = {0: "月", 1: "火", 2: "水", 3: "木", 4: "金", 5: "土", 6: "日"}
        return wn_wd[wd]


def resource_path(relative):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative)
    return os.path.join(os.path.abspath('.'), relative)

def main():
    app = QApplication(sys.argv)
    # app.setWindowIcon(QIcon(resource_path('logo.png')))
    # window = Splash()
    MainWindow = Application()
    MainWindow.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()